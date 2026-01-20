import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from frappe.utils.file_manager import save_file
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO

@frappe.whitelist()
def export_to_xf_excel(
    doctype, doc_name, excel_format=None, attach=False, download=False,is_preview=False
):

    doc = frappe.get_doc(doctype, doc_name)

    excelFormat = None
    if excel_format:

        excelFormat = frappe.get_value(
            "XF Excel Format",
            {"doc_type": doctype, "name": excel_format, "enable": 1},
            "python_script",
        )
    else:

        excelFormat = frappe.get_value(
            "XF Excel Format",
            {"doc_type": doctype, "is_default": 1, "enable": 1},
            "python_script  ",
        )
        if not excelFormat:
            frappe.throw("Default enabled Excel Format not defined for this Doctype")

    if not excelFormat:
        frappe.throw("Enabled Excel Format not defined for this Doctype")

    wb = openpyxl.Workbook()
    ws = wb.active
    context = {
        "wb": wb,
        "ws": ws,
        "Font": Font,
        "doc": doc,
        "Alignment": Alignment,
        "get_column_letter": get_column_letter,
        "Border": Border,
        "Side": Side,
        "PatternFill": PatternFill,
        "Image": Image,
    }
    
    exec(excelFormat, context)
    
        
    ws = context["ws"]

    from io import BytesIO

    output = BytesIO()
    context["wb"].save(output)
    output.seek(0)

    file_name = f"XF_Excel_{doctype}_{doc_name}.xlsx"

    if is_preview:
        html_preview = generate_excel_preview(output)
        return {"preview": html_preview}

    if attach:
        file_doc = save_file(file_name, output.read(), doctype, doc_name, is_private=1)
        return file_doc.file_url

    if download:
        frappe.local.response.filename = file_name
        frappe.local.response.filecontent = output.getvalue()
        frappe.local.response.type = "binary"
        frappe.local.response.mime_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return


@frappe.whitelist()
def get_default_excel_format(doctype):
    default_format = frappe.db.get_value(
        "XF Excel Format", {"doc_type": doctype, "is_default": 1}, "name"
    )
    return default_format


def rgb_to_hex(rgb):
    """Convert Excel's ARGB color to HEX"""
    if rgb is None or rgb == "00000000":
        return "#FFFFFF"  
    return f"#{rgb[-6:]}"

def generate_excel_preview(output):
    output.seek(0)
    wb = load_workbook(BytesIO(output.read()))
    ws = wb.active  

    merged_cells = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        merged_cells[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)

  
    html = "<table border='1' style='border-collapse: collapse; width: 100%; text-align: center;'>"

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        html += "<tr>"
        for col_idx, cell in enumerate(row, start=1):
            if any((row_idx, col_idx) in merged_range for merged_range in merged_cells if (row_idx, col_idx) != merged_range):
                continue  

            rowspan, colspan = merged_cells.get((row_idx, col_idx), (1, 1))
            rowspan_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ""
            colspan_attr = f' colspan="{colspan}"' if colspan > 1 else ""

           
            bold = "font-weight: bold;" if cell.font and cell.font.bold else ""
            bg_color = rgb_to_hex(cell.fill.start_color.rgb)  
            text_align = "center" 
            if cell.alignment:
                if cell.alignment.horizontal:
                    text_align = cell.alignment.horizontal
                elif isinstance(cell.alignment, Alignment):
                    text_align = cell.alignment.horizontal or "center"

           
            value = cell.value if cell.value is not None else ""

            cell_style = f"background-color: {bg_color}; text-align: {text_align}; padding: 5px; border: 1px solid black; {bold}".strip()
            html += f"<td style=\"{cell_style}\"{rowspan_attr}{colspan_attr}>{value}</td>"
        html += "</tr>"

    html += "</table>"

    return html

